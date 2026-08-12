from __future__ import annotations
import re
from pathlib import Path
import openpyxl

CATEGORIES = ['1','2','3','4','5A','5B','5C','5D','6','7A','7B','8','9','10A','10B','11A','11B','12']


def _numeric_limit(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    m = re.search(r'[-+]?\d+(?:\.\d+)?(?:[Ee][-+]?\d+)?', s.replace(',', '.'))
    return float(m.group()) if m else None


def import_ifra_overview(db, xlsx_path: str | Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = ws.iter_rows(values_only=True)
    try:
        next(all_rows); next(all_rows); header_values = next(all_rows)
    except StopIteration:
        raise ValueError('IFRA overview workbook is too short')
    headers = {str(v).strip(): i for i, v in enumerate(header_values) if v is not None}
    required = ['Key','Amendment number','Name of the IFRA Standard','CAS numbers','IFRA Standard type']
    missing = [x for x in required if x not in headers]
    if missing:
        raise ValueError(f'Not recognized as IFRA overview. Missing: {missing}')

    def val(row, header, default=''):
        idx=headers.get(header)
        if idx is None or idx >= len(row): return default
        v=row[idx]
        return default if v is None else v

    with db.connect() as conn:
        conn.execute('DELETE FROM ifra_limits')
        conn.execute('DELETE FROM ifra_standards')
        count = 0
        limits_count = 0
        for row in all_rows:
            key = val(row,'Key',None); name = val(row,'Name of the IFRA Standard',None)
            if not key or not name: continue
            amendment = val(row,'Amendment number',None)
            notes_parts=[]
            for h in ['Prohibited fragrance ingredients: notes','Restricted ingredients: notes','Specified ingredients: notes','Contributions from other sources: notes']:
                v=val(row,h,'')
                if v: notes_parts.append(str(v))
            cur=conn.execute('''INSERT INTO ifra_standards(std_key, amendment, name, cas_numbers, synonyms, standard_type, risk_property, notes)
                                VALUES(?,?,?,?,?,?,?,?)''',(
                str(key), int(amendment) if amendment is not None else None, str(name), str(val(row,'CAS numbers','')),
                str(val(row,'Synonyms','')), str(val(row,'IFRA Standard type','')), str(val(row,'Intrinsic property driving the risk management measure','')), '\n'.join(notes_parts)))
            std_id=cur.lastrowid; count+=1
            for cat in CATEGORIES:
                raw=val(row,f'Category {cat} (%)',None); limit=_numeric_limit(raw)
                conn.execute('INSERT INTO ifra_limits(standard_id,category,max_pct,raw_value) VALUES(?,?,?,?)',
                             (std_id,cat,limit,'' if raw is None else str(raw)))
                limits_count+=1
    wb.close()
    return {'standards':count,'limits':limits_count}


def import_ncs_annex(db, xlsx_path: str | Path):
    wb=openpyxl.load_workbook(xlsx_path,data_only=True,read_only=True)
    if 'Natural contributions' not in wb.sheetnames:
        raise ValueError('Natural contributions sheet not found')
    ws=wb['Natural contributions']
    rows=ws.iter_rows(values_only=True)
    header_values=None
    for i,row in enumerate(rows,1):
        if i>20: break
        if 'NCS NAME' in row and 'CAS NUMBER CONSTITUENT (IFRA STANDARD)' in row:
            header_values=row; break
    if not header_values: raise ValueError('Could not find NCS header row')
    headers={str(v).strip():i for i,v in enumerate(header_values) if v is not None}
    def val(row, header, default=''):
        idx=headers.get(header.strip())
        if idx is None or idx>=len(row): return default
        v=row[idx]; return default if v is None else v

    with db.connect() as conn:
        conn.execute('DELETE FROM ncs_contributions')
        count=0
        for row in rows:
            ncs_name=val(row,'NCS NAME',None); constituent_cas=val(row,'CAS NUMBER CONSTITUENT (IFRA STANDARD)',None); conc=val(row,'CONCENTRATION OF CONSTITUENT IN NATURALS (NCS) (%)',None)
            if not ncs_name or not constituent_cas or conc is None: continue
            try: conc=float(conc)
            except Exception: continue
            amendment=val(row,'IFRA Amendment (review)',51)
            try: amendment=int(amendment)
            except Exception: amendment=51
            cur=conn.execute('''INSERT OR IGNORE INTO ncs_contributions(amendment,ncs_name,botanical_name,principal_cas,other_cas,constituent_name,constituent_cas,concentration_pct)
                            VALUES(?,?,?,?,?,?,?,?)''',(
                amendment,str(ncs_name),str(val(row,'NCS BOTANICAL NAME','')),str(val(row,'PRINCIPAL CAS NUMBER OF NCS','')),str(val(row,'OTHER CAS NUMBERS OF NCS','')),
                str(val(row,'CONSTITUENT NAME (IFRA STANDARD)','')),str(constituent_cas),conc))
            if cur.rowcount: count+=1
    wb.close()
    return {'ncs_contributions':count}
